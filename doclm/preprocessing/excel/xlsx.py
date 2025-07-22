
import yaml
import logging

import numpy as np
import pandas as pd
import hnswlib

from sklearn.preprocessing import MultiLabelBinarizer

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..common import get_headers, get_file_info_from_reader


log = logging.getLogger("doclogger")


def xlsx_processor(stream, **kwargs):
    """Partitions Microsoft Excel Documents in .xlsx format into its document elements.
    """

    wb = load_workbook(stream, data_only=True)

    meta_info = get_file_info_from_reader(stream)
    sheets = {sheet_name: wb[sheet_name]
              for sheet_name in wb.sheetnames if isinstance(wb[sheet_name], Worksheet)}
    meta_info['pages'] = len(sheets)
    return meta_info, extract_sheets(sheets)


def extract_sheets(workbook_sheets:dict):
    parser = ExcelParser(eps=10, min_samples=3)

    for sheet_name, sheet in workbook_sheets.items():
        # sheet = workbook[sheet_name]
        tables = parser.extract_tables(sheet)
        yield sheet_name, yaml.dump(list(tables))


class ExcelParser:
    """
    using dbscan with HNSW for distance search
    """
    def __init__(self, eps, min_samples, space='l2'):

        self.eps = eps
        self.min_samples = min_samples
        self.space = space
        self.mlb = MultiLabelBinarizer()

    @staticmethod
    def get_cell_color(cell):
        color = cell.fill.start_color.index
        if not isinstance(color, int):
            return str(int(color[2:], 16))
        return str(int(color))


    @staticmethod
    def getCellBorders(cell_ref):
        tmp = cell_ref.border
        brdrs = []

        if tmp.top.style is not None: brdrs.append('T')
        if tmp.left.style is not None: brdrs.append('L')
        if tmp.right.style is not None: brdrs.append('R')
        if tmp.bottom.style is not None: brdrs.append('B')
        return brdrs


    @staticmethod
    def getCellAlignment(cell_ref):
        tmp = cell_ref.alignment
        almnt = []

        if tmp.indent is not None: almnt.append(str(tmp.indent))
        if tmp.horizontal is not None: almnt.append('H')
        if tmp.vertical is not None: almnt.append('V')
        if tmp.textRotation is not None: almnt.append('R')
        if tmp.wrapText is not None: almnt.append('W')
        if tmp.shrinkToFit is not None: almnt.append('S')

        return almnt


    def extract_data(self, sheet_obj):
        data_points = []
        value = []
        for row in sheet_obj.iter_rows():
            for cell_obj in row:

                if cell_obj.value is None:
                    continue

                data_points.append(
                    [cell_obj.data_type, self.get_cell_color(cell_obj), str(cell_obj.is_date), cell_obj.style,
                     cell_obj.font.name,] + self.getCellAlignment(cell_obj) + self.getCellBorders(cell_obj)
                   )
                value.append([cell_obj.row, cell_obj.column,  cell_obj.value])
        cols = ['dtype', 'color', 'is_date', 'style', 'font', 'alinment', 'border']
        return cols, data_points, pd.DataFrame(value)

    # pd.DataFrame(coordinates_value_[labels == i])
    def make_clusters(self, data):

        noise = -1
        dim = data.shape[1]
        num_elements = data.shape[0]

        hnsw_index = hnswlib.Index(space=self.space, dim=dim)
        hnsw_index.init_index(max_elements=num_elements, ef_construction=200, M=16)
        hnsw_index.add_items(data)
        hnsw_index.set_ef(100)

        def hnsw_neighbors(point, eps):
            k = min(50, num_elements)
            n_labels, distances = hnsw_index.knn_query(point, k=k)
            return n_labels[0][distances[0] <= eps]

    # Custom DBSCAN using HNSW for neighbor search
        cluster_id = 0
        labels = np.full(num_elements, noise)

        for i in range(num_elements):
            if labels[i] != noise:
                continue
            neighbors = hnsw_neighbors(data[i], self.eps)

            if len(neighbors) > self.min_samples:                   # check if core point
                labels[i] = cluster_id
                neighbors = list(neighbors)
                while neighbors:
                    neighbor_point = neighbors.pop()                # get an element from the possible cluster

                    if labels[neighbor_point] == noise:
                        labels[neighbor_point] = cluster_id

                        new_neighbors = hnsw_neighbors(data[neighbor_point], self.eps)
                        if len(new_neighbors) >= self.min_samples:   # check if it is also a core point
                            neighbors.extend(new_neighbors)          #  expand the cluster from the current point

            cluster_id += 1
        return labels

    def extract_tables(self, sh):

        columns, data_features, coordinates_value_ = self.extract_data(sh)
        binarized_data = self.mlb.fit_transform(data_features)
        data_for_cluster = np.concatenate([coordinates_value_[[0, 1]].values,
                                           binarized_data], axis=1)
        labels = self.make_clusters(data_for_cluster)
        log.debug('clusters label found %s ', np.unique(labels))
        for i in np.unique(labels):
            sparse_cluster_points = coordinates_value_[labels==i]
            # if i == -1: # the noise part
            #     pass
            table = sparse_cluster_points.pivot_table(index=0, columns=1, values=2, aggfunc=lambda x: x)
            yield self.process_pandas_table(table)

    @staticmethod
    def process_pandas_table(p_df):
        header_indices, cols = get_headers(p_df.iloc[:10])
        p_df.columns = cols
        p_df.drop(index=header_indices, inplace=True)
        return p_df.to_csv(na_rep="", index=None)
    # numerics = table[1][table[1].str.isnumeric()]